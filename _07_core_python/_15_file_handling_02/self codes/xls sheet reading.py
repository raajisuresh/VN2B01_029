import openpyxl
file_path='D:\\Users\\skred\\Desktop\\MKR\\MKR EXPENSES&SUMMARY SHEET.xlsx'
wb_obj=openpyxl.load_workbook(file_path)
sheet_obj=wb_obj.active
print(sheet_obj.max_row)
print(sheet_obj.active_cell)
max=sheet_obj.max_column
max_r=sheet_obj.max_row
print(max)
for i in range(4,max+1):
    for j in range(2,max_r):
        cell_obj=sheet_obj.cell(row=j,column=i)
        print(cell_obj.value)